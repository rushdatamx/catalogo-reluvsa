"""
Router del carrito persistido en el SERVIDOR.

¿Por qué existe?
  El carrito vivía solo en el localStorage del navegador. Si el proveedor armaba
  un pedido grande y no alcanzaba a enviarlo (o limpiaba caché / cambiaba de
  dispositivo), ese trabajo se perdía SIN RASTRO: nunca llegaba al servidor, así
  que no había forma de recuperarlo ni de que el admin lo viera.

  Caso real (julio 2026): un proveedor armó 141 productos, no le dio a "Enviar
  pedido" y el carrito desapareció. Irrecuperable.

Qué resuelve:
  1. El carrito se guarda en `pedidos.db` (volumen persistente) en cada cambio.
     Se recupera desde cualquier dispositivo y sobrevive a limpiar el navegador.
  2. El admin puede ver los carritos ACTIVOS (armados pero no enviados) y
     rescatarlos como CSV — si un proveedor se atora, RELUVSA ya sabe qué quería.
  3. Importación masiva por SKU (Excel/CSV): armar 141 renglones a mano en la UI
     es inviable; las refaccionarias ya tienen su lista en archivo.

Notas de diseño:
  - Se guardan SOLO {sku, cantidad}. Precio, nombre e inventario se rehidratan
    del catálogo al leer, porque cambian con cada actualización mensual y no
    queremos servir precios rancios desde un carrito viejo.
  - Guardar es idempotente y "last write wins": un carrito por usuario.
  - Los SKUs se normalizan igual que en el catálogo (quitando ceros iniciales en
    numéricos) porque el Excel del cliente los trae con ceros a la izquierda.
"""
import csv
import io
import json
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel, Field

from database import get_db
from pedidos_db import get_pedidos_db
from routers.auth import get_current_user, require_admin, UserInfo

router = APIRouter(prefix="/api", tags=["carrito"])

# Tope de renglones distintos por carrito. Alto a propósito (pedidos reales de
# refaccionaria pasan de 140), pero acotado para no aceptar cargas absurdas.
MAX_RENGLONES = 1000
# Cantidad máxima por renglón (coincide con CarritoItem de pagos.py).
MAX_CANTIDAD = 999


# --- Modelos ----------------------------------------------------------------
class CarritoItemIn(BaseModel):
    sku: str
    cantidad: int = Field(gt=0, le=MAX_CANTIDAD)


class CarritoIn(BaseModel):
    items: List[CarritoItemIn] = Field(default_factory=list)


class LineaImportada(BaseModel):
    sku: str
    cantidad: int
    nombre: Optional[str] = None
    marca: Optional[str] = None
    precio: Optional[float] = None
    inventario: Optional[int] = None


class ResultadoImportacion(BaseModel):
    """Resultado de importar un archivo: qué entró y qué NO, con el motivo.

    Nunca se falla en bloque: el proveedor se queda con lo que sí sirve y ve
    exactamente qué renglones revisar."""
    items: List[LineaImportada]
    encontrados: int
    no_encontrados: List[str]
    sin_stock: List[str]
    ajustados: List[str]          # se pidió más de lo disponible -> se bajó al stock
    filas_ignoradas: int          # filas sin SKU legible (encabezados, vacías)


# --- Helpers ----------------------------------------------------------------
def _normalizar_sku(sku: str) -> str:
    """Normaliza un SKU para hacer match con el catálogo.

    El Excel del cliente trae SKUs numéricos con ceros a la izquierda
    ('013030102') que en BD viven sin ellos ('13030102'). Solo se quitan si el
    valor es puramente numérico: los alfanuméricos se dejan intactos."""
    s = (sku or "").strip().upper()
    if s.isdigit():
        return s.lstrip("0") or "0"
    return s


def _hidratar(items: list, conn) -> tuple[list, list]:
    """Completa {sku, cantidad} con datos vivos del catálogo.

    Devuelve (items_hidratados, skus_no_encontrados). Los precios/inventario se
    leen SIEMPRE del catálogo (no se confía en lo guardado) porque cambian con
    cada actualización mensual."""
    if not items:
        return [], []

    cursor = conn.cursor()
    hidratados: list = []
    faltantes: list = []

    for it in items:
        sku = it["sku"]
        cursor.execute(
            """SELECT sku, nombre_producto, tipo_producto, marca,
                      precio_publico, inventario_total
               FROM productos WHERE sku = ?""",
            (sku,),
        )
        prod = cursor.fetchone()
        if prod is None:
            faltantes.append(sku)
            continue
        hidratados.append({
            "sku": prod["sku"],
            "nombre": prod["nombre_producto"] or prod["tipo_producto"] or prod["sku"],
            "marca": prod["marca"],
            "precio": prod["precio_publico"] or 0,
            "inventario": prod["inventario_total"] or 0,
            "cantidad": it["cantidad"],
        })

    return hidratados, faltantes


def _leer_filas(contenido: bytes, filename: str) -> list[tuple[str, str]]:
    """Extrae (sku, cantidad) crudos de un CSV o XLSX.

    Heurística deliberadamente permisiva: los archivos del cliente no tienen un
    formato fijo. Se busca por nombre de encabezado y, si no hay encabezado
    reconocible, se cae a 'primera columna = SKU, segunda = cantidad'."""
    nombre = (filename or "").lower()
    filas: list[list] = []

    if nombre.endswith(".xlsx") or nombre.endswith(".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            filas.append(list(row))
        wb.close()
    else:
        # CSV/TXT. utf-8-sig se come el BOM que mete Excel al exportar.
        try:
            texto = contenido.decode("utf-8-sig")
        except UnicodeDecodeError:
            texto = contenido.decode("latin-1")
        # Detectar el separador (coma, punto y coma o tab) de forma tolerante.
        muestra = texto[:4096]
        try:
            dialecto = csv.Sniffer().sniff(muestra, delimiters=",;\t|")
            sep = dialecto.delimiter
        except csv.Error:
            sep = ","
        filas = [r for r in csv.reader(io.StringIO(texto), delimiter=sep)]

    if not filas:
        return []

    # ¿La primera fila es encabezado? Buscar columnas por nombre.
    idx_sku, idx_cant = 0, 1
    encabezado = [str(c or "").strip().lower() for c in filas[0]]
    CLAVES_SKU = ("sku", "clave", "codigo", "código", "articulo", "artículo", "parte")
    CLAVES_CANT = ("cantidad", "cant", "qty", "piezas", "pzas", "pz")

    tiene_encabezado = False
    for i, celda in enumerate(encabezado):
        if any(k in celda for k in CLAVES_SKU):
            idx_sku, tiene_encabezado = i, True
        elif any(k in celda for k in CLAVES_CANT):
            idx_cant, tiene_encabezado = i, True

    cuerpo = filas[1:] if tiene_encabezado else filas
    resultado: list[tuple[str, str]] = []
    for fila in cuerpo:
        if not fila:
            continue
        sku = str(fila[idx_sku]).strip() if idx_sku < len(fila) and fila[idx_sku] is not None else ""
        cant = str(fila[idx_cant]).strip() if idx_cant < len(fila) and fila[idx_cant] is not None else ""
        resultado.append((sku, cant))
    return resultado


# --- Endpoints: carrito del usuario -----------------------------------------
@router.get("/carrito")
def obtener_carrito(user: UserInfo = Depends(get_current_user)):
    """Carrito guardado del usuario, hidratado con precios/inventario actuales.

    Si un producto guardado ya no existe en el catálogo, se reporta en
    `removidos` en vez de romper la carga."""
    with get_pedidos_db() as pconn:
        row = pconn.execute(
            "SELECT items, updated_at FROM carritos WHERE username = ?",
            (user.username,),
        ).fetchone()

    if row is None:
        return {"items": [], "updated_at": None, "removidos": []}

    try:
        guardados = json.loads(row["items"]) or []
    except (json.JSONDecodeError, TypeError):
        guardados = []

    with get_db() as conn:
        items, faltantes = _hidratar(guardados, conn)

    return {"items": items, "updated_at": row["updated_at"], "removidos": faltantes}


@router.put("/carrito")
def guardar_carrito(data: CarritoIn, user: UserInfo = Depends(get_current_user)):
    """Guarda (sobrescribe) el carrito del usuario. Last write wins.

    Solo persiste {sku, cantidad}: el resto se rehidrata al leer."""
    if len(data.items) > MAX_RENGLONES:
        raise HTTPException(
            status_code=400,
            detail=f"El carrito supera el máximo de {MAX_RENGLONES} renglones distintos",
        )

    # Consolidar por SKU (el mismo producto no debe ocupar dos renglones).
    cantidades: dict[str, int] = {}
    for it in data.items:
        sku = it.sku.strip()
        if not sku:
            continue
        cantidades[sku] = min(cantidades.get(sku, 0) + it.cantidad, MAX_CANTIDAD)

    payload = json.dumps([{"sku": s, "cantidad": c} for s, c in cantidades.items()])

    with get_pedidos_db() as pconn:
        pconn.execute(
            """INSERT INTO carritos (username, items, updated_at)
               VALUES (?, ?, CURRENT_TIMESTAMP)
               ON CONFLICT(username) DO UPDATE
                 SET items = excluded.items, updated_at = CURRENT_TIMESTAMP""",
            (user.username, payload),
        )
    return {"guardado": True, "renglones": len(cantidades)}


@router.delete("/carrito")
def vaciar_carrito(user: UserInfo = Depends(get_current_user)):
    """Vacía el carrito guardado (se llama al enviar el pedido)."""
    with get_pedidos_db() as pconn:
        pconn.execute("DELETE FROM carritos WHERE username = ?", (user.username,))
    return {"vaciado": True}


@router.post("/carrito/importar", response_model=ResultadoImportacion)
async def importar_carrito(
    archivo: UploadFile = File(...),
    user: UserInfo = Depends(get_current_user),
):
    """Convierte un Excel/CSV de (SKU, cantidad) en renglones de carrito.

    NO guarda nada: devuelve lo que encontró para que el frontend lo muestre y
    el usuario confirme. Así un archivo con errores no pisa un carrito bueno.

    Tolerante por diseño: reporta los renglones problemáticos en vez de fallar
    en bloque — con listas de 141 productos, abortar por un SKU malo es inútil.
    """
    contenido = await archivo.read()
    if not contenido:
        raise HTTPException(status_code=400, detail="El archivo está vacío")
    # 5 MB es holgado para una lista de SKUs y evita cargas accidentales enormes.
    if len(contenido) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El archivo excede 5 MB")

    try:
        crudas = _leer_filas(contenido, archivo.filename or "")
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="No se pudo leer el archivo. Debe ser .xlsx o .csv con SKU y cantidad.",
        )

    if not crudas:
        raise HTTPException(status_code=400, detail="El archivo no tiene filas legibles")

    # Consolidar por SKU normalizado, ignorando filas sin SKU.
    cantidades: dict[str, int] = {}
    ignoradas = 0
    for sku_crudo, cant_cruda in crudas:
        sku = _normalizar_sku(sku_crudo)
        if not sku:
            ignoradas += 1
            continue
        try:
            # float() primero: Excel entrega "3.0" para un 3.
            cantidad = int(float(str(cant_cruda).replace(",", "").strip() or 1))
        except (ValueError, TypeError):
            cantidad = 1
        if cantidad <= 0:
            ignoradas += 1
            continue
        cantidades[sku] = min(cantidades.get(sku, 0) + cantidad, MAX_CANTIDAD)

    if not cantidades:
        raise HTTPException(
            status_code=400,
            detail="No se encontró ninguna columna con SKUs válidos en el archivo",
        )

    if len(cantidades) > MAX_RENGLONES:
        raise HTTPException(
            status_code=400,
            detail=f"El archivo trae {len(cantidades)} SKUs distintos (máximo {MAX_RENGLONES})",
        )

    items: list = []
    no_encontrados: list = []
    sin_stock: list = []
    ajustados: list = []

    with get_db() as conn:
        cursor = conn.cursor()
        for sku, cantidad in cantidades.items():
            cursor.execute(
                """SELECT sku, nombre_producto, tipo_producto, marca,
                          precio_publico, inventario_total
                   FROM productos WHERE sku = ?""",
                (sku,),
            )
            prod = cursor.fetchone()
            if prod is None:
                no_encontrados.append(sku)
                continue

            inventario = prod["inventario_total"] or 0
            precio = prod["precio_publico"] or 0
            # Misma regla AGOTADO que el checkout: sin precio o sin stock no se pide.
            if inventario <= 0 or precio <= 0:
                sin_stock.append(prod["sku"])
                continue

            final = cantidad
            if cantidad > inventario:
                final = inventario
                ajustados.append(f"{prod['sku']}: {cantidad}→{int(inventario)}")

            items.append(LineaImportada(
                sku=prod["sku"],
                cantidad=final,
                nombre=prod["nombre_producto"] or prod["tipo_producto"] or prod["sku"],
                marca=prod["marca"],
                precio=precio,
                inventario=inventario,
            ))

    # Si NADA hizo match, casi siempre es que se subió el archivo equivocado o
    # que la columna de SKU no es la que se detectó. Devolver "0 encontrados"
    # con 200 deja al usuario sin saber si su archivo está mal o si sus SKUs no
    # existen; un error explícito le dice qué revisar.
    if not items and not sin_stock:
        muestra = ", ".join(no_encontrados[:5])
        raise HTTPException(
            status_code=400,
            detail=(
                f"Ninguno de los {len(no_encontrados)} códigos del archivo existe en "
                f"el catálogo (ej.: {muestra}). Revisa que la columna de SKU sea la "
                "correcta y que sean claves de RELUVSA."
            ),
        )

    return ResultadoImportacion(
        items=items,
        encontrados=len(items),
        no_encontrados=no_encontrados,
        sin_stock=sin_stock,
        ajustados=ajustados,
        filas_ignoradas=ignoradas,
    )


# --- Endpoints: vista de admin ----------------------------------------------
@router.get("/carritos")
def listar_carritos(admin: UserInfo = Depends(require_admin)):
    """Carritos ACTIVOS: armados pero todavía no enviados como pedido.

    Es la red de seguridad del caso que originó esta feature — si un proveedor
    arma un pedido grande y se atora, RELUVSA lo ve aquí y puede levantarlo por
    teléfono en vez de perderlo."""
    with get_pedidos_db() as pconn:
        filas = pconn.execute(
            """SELECT c.username, u.nombre_empresa, u.contacto, c.items, c.updated_at
               FROM carritos c
               LEFT JOIN usuarios u ON u.username = c.username
               ORDER BY c.updated_at DESC"""
        ).fetchall()

    carritos: list = []
    with get_db() as conn:
        for f in filas:
            try:
                guardados = json.loads(f["items"]) or []
            except (json.JSONDecodeError, TypeError):
                guardados = []
            if not guardados:
                continue  # carrito vacío: no es nada que rescatar
            items, _ = _hidratar(guardados, conn)
            carritos.append({
                "username": f["username"],
                "nombre_empresa": f["nombre_empresa"],
                "contacto": f["contacto"],
                "updated_at": f["updated_at"],
                "num_renglones": len(items),
                "num_piezas": sum(i["cantidad"] for i in items),
                "total_estimado": sum(i["precio"] * i["cantidad"] for i in items),
                "items": items,
            })
    return {"carritos": carritos}
