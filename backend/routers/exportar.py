"""
Endpoints para exportar catálogo a Excel y PDF con imágenes embebidas.
"""
import asyncio
import io
import json
import math
import sqlite3
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse
import httpx
from PIL import Image as PILImage

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
)

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from database import get_db
from utils.busqueda_inteligente import analizar_busqueda, STOP_WORDS
from utils.sinonimos import expandir_sinonimos

router = APIRouter(prefix="/api/exportar", tags=["exportar"])

# URL base del servidor de imágenes
IMAGE_SERVER_BASE = "http://reluvsa.zapto.org/cgi-vel/omner2/fotos"

MAX_PRODUCTOS_EXPORT = 500
IMAGE_DOWNLOAD_TIMEOUT = 5.0
IMAGE_CONCURRENCY = 10
THUMB_SIZE_EXCEL = (75, 75)
THUMB_SIZE_PDF = (150, 150)  # 3x resolution, rendered at 50x50pt for sharp output


# ---------------------------------------------------------------------------
# Shared: Build query (replicates productos.py filter logic)
# ---------------------------------------------------------------------------

def _build_filtered_query(
    departamento, marca, marca_vehiculo, modelo_vehiculo, año, motor,
    tipo_producto, grupo_producto, con_inventario, solo_nuevos,
    ancho_llanta, relacion_llanta, diametro_llanta, tipo_llanta, capas_llanta,
    viscosidad, tipo_aceite, presentacion,
    grupo_bci, capacidad_cca, tamano_acumulador,
    q,
):
    """Returns (select_sql, params) for the filtered product list (no pagination)."""

    base_query = "FROM productos p"
    where_clauses: List[str] = []
    params: List[Any] = []

    # Conditional JOINs
    needs_compat_join = any([marca_vehiculo, modelo_vehiculo, año, motor])
    if needs_compat_join:
        base_query += " INNER JOIN compatibilidades c ON c.producto_id = p.id"

    needs_carac_join = any([
        ancho_llanta, relacion_llanta, diametro_llanta, tipo_llanta, capas_llanta,
        viscosidad, tipo_aceite, presentacion,
        grupo_bci, capacidad_cca, tamano_acumulador,
    ])
    if needs_carac_join:
        base_query += " INNER JOIN caracteristicas_producto cp ON cp.producto_id = p.id"

    # Basic filters
    if departamento:
        where_clauses.append("p.departamento = ?")
        params.append(departamento)
    if marca:
        where_clauses.append("p.marca = ?")
        params.append(marca)
    if tipo_producto:
        where_clauses.append("p.tipo_producto = ?")
        params.append(tipo_producto)
    if grupo_producto:
        where_clauses.append("p.grupo_producto = ?")
        params.append(grupo_producto)
    if con_inventario:
        where_clauses.append("p.inventario_total > 0")
    if solo_nuevos:
        where_clauses.append("p.created_at >= datetime('now', '-60 days')")

    # Vehicle filters
    if marca_vehiculo:
        where_clauses.append("c.marca_vehiculo = ?")
        params.append(marca_vehiculo)
    if modelo_vehiculo:
        where_clauses.append("c.modelo_vehiculo = ?")
        params.append(modelo_vehiculo)
    if año:
        where_clauses.append("c.año_inicio <= ? AND c.año_fin >= ?")
        params.extend([año, año])
    if motor:
        where_clauses.append("c.motor = ?")
        params.append(motor)

    # Smart search
    if q:
        analisis = analizar_busqueda(q)

        if analisis["tipo"] in ("combinada", "solo_vehiculo"):
            if "INNER JOIN compatibilidades" not in base_query:
                base_query += " INNER JOIN compatibilidades c ON c.producto_id = p.id"

            vehiculo = analisis['vehiculo'].upper()
            if analisis["tipo_vehiculo"] == "modelo":
                where_clauses.append("c.modelo_vehiculo = ?")
            else:
                where_clauses.append("c.marca_vehiculo = ?")
            params.append(vehiculo)

            if analisis.get("año"):
                where_clauses.append("c.año_inicio <= ? AND c.año_fin >= ?")
                params.extend([analisis["año"], analisis["año"]])

            if analisis["tipo"] == "combinada":
                palabras_producto = analisis['producto'].split()
                grupos_sin = expandir_sinonimos(palabras_producto)
                condiciones_grupos = []
                for grupo in grupos_sin:
                    or_parts = []
                    for termino in grupo:
                        term = f"%{termino}%"
                        or_parts.append(
                            "(p.nombre_producto LIKE ? OR p.grupo_producto LIKE ? "
                            "OR p.tipo_producto LIKE ? OR p.descripcion_original LIKE ?)"
                        )
                        params.extend([term] * 4)
                    condiciones_grupos.append(f"({' OR '.join(or_parts)})")
                if condiciones_grupos:
                    where_clauses.append(f"({' AND '.join(condiciones_grupos)})")
        else:
            palabras = [p.lower() for p in q.strip().split() if p.lower() not in STOP_WORDS]
            if not palabras:
                palabras = [p.lower() for p in q.strip().split()]

            grupos_sin = expandir_sinonimos(palabras)

            if len(grupos_sin) == 1 and len(grupos_sin[0]) == 1:
                search_term = f"%{q}%"
                where_clauses.append("""
                    (
                        p.descripcion_original LIKE ?
                        OR p.sku LIKE ?
                        OR p.nombre_producto LIKE ?
                        OR p.skus_alternos LIKE ?
                        OR p.marca LIKE ?
                        OR p.departamento LIKE ?
                        OR p.tipo_producto LIKE ?
                        OR p.id IN (
                            SELECT producto_id FROM compatibilidades
                            WHERE marca_vehiculo LIKE ?
                               OR modelo_vehiculo LIKE ?
                               OR motor LIKE ?
                        )
                    )
                """)
                params.extend([search_term] * 10)
            else:
                condiciones_grupos = []
                for grupo in grupos_sin:
                    or_parts = []
                    for termino in grupo:
                        term = f"%{termino}%"
                        or_parts.append(
                            "(p.nombre_producto LIKE ? OR p.tipo_producto LIKE ? "
                            "OR p.descripcion_original LIKE ? OR p.grupo_producto LIKE ?)"
                        )
                        params.extend([term] * 4)

                    palabras_originales = [t for t in grupo if t in palabras or t in q.lower()]
                    for orig in palabras_originales:
                        orig_term = f"%{orig}%"
                        or_parts.append(
                            "(p.sku LIKE ? OR p.skus_alternos LIKE ? "
                            "OR p.marca LIKE ? OR p.departamento LIKE ?)"
                        )
                        params.extend([orig_term] * 4)

                    condiciones_grupos.append(f"({' OR '.join(or_parts)})")
                where_clauses.append(f"({' AND '.join(condiciones_grupos)})")

    # Characteristic subquery filters (tires, oils, batteries)
    _char_filters = [
        (ancho_llanta, 'llanta', 'ancho'),
        (relacion_llanta, 'llanta', 'relacion'),
        (diametro_llanta, 'llanta', 'diametro'),
        (tipo_llanta, 'llanta', 'tipo'),
        (capas_llanta, 'llanta', 'capas'),
        (viscosidad, 'aceite', 'viscosidad'),
        (tipo_aceite, 'aceite', 'tipo_aceite'),
        (presentacion, 'aceite', 'presentacion'),
        (grupo_bci, 'acumulador', 'grupo_bci'),
        (capacidad_cca, 'acumulador', 'capacidad_cca'),
        (tamano_acumulador, 'acumulador', 'tamano'),
    ]
    for valor, cat, clave in _char_filters:
        if valor:
            where_clauses.append(
                f"p.id IN (SELECT producto_id FROM caracteristicas_producto "
                f"WHERE categoria = ? AND clave = ? AND valor = ?)"
            )
            params.extend([cat, clave, valor])

    where_sql = ""
    if where_clauses:
        where_sql = "WHERE " + " AND ".join(where_clauses)

    select_sql = f"""
        SELECT DISTINCT
            p.id, p.sku, p.departamento, p.marca, p.descripcion_original,
            p.nombre_producto, p.tipo_producto, p.precio_publico, p.precio_mayoreo,
            p.imagen_url, p.inventario_total, p.grupo_producto, p.skus_alternos,
            CASE WHEN p.created_at >= datetime('now', '-60 days') THEN 1 ELSE 0 END as es_nuevo
        {base_query}
        {where_sql}
        ORDER BY p.inventario_total DESC, p.sku
    """
    return select_sql, params


# ---------------------------------------------------------------------------
# Shared: Fetch enrichment data in bulk
# ---------------------------------------------------------------------------

def _fetch_compatibilidades_bulk(conn, product_ids: List[int]) -> Dict[int, List[str]]:
    """Returns {product_id: [formatted_compat_strings]}."""
    if not product_ids:
        return {}
    placeholders = ",".join("?" * len(product_ids))
    cursor = conn.cursor()
    cursor.execute(f"""
        SELECT producto_id, marca_vehiculo, modelo_vehiculo, año_inicio, año_fin, motor
        FROM compatibilidades
        WHERE producto_id IN ({placeholders})
        ORDER BY producto_id, marca_vehiculo, modelo_vehiculo
    """, product_ids)

    result: Dict[int, List[str]] = {}
    for row in cursor.fetchall():
        pid = row['producto_id']
        parts = []
        if row['marca_vehiculo']:
            parts.append(row['marca_vehiculo'])
        if row['modelo_vehiculo']:
            parts.append(row['modelo_vehiculo'])
        if row['año_inicio'] and row['año_fin']:
            parts.append(f"{row['año_inicio']}-{row['año_fin']}")
        if row['motor']:
            parts.append(row['motor'])
        line = " ".join(parts)
        result.setdefault(pid, []).append(line)
    return result


def _fetch_intercambiables_bulk(conn, product_ids: List[int]) -> Dict[int, List[str]]:
    """Returns {product_id: [formatted_interchange_strings]}."""
    if not product_ids:
        return {}
    placeholders = ",".join("?" * len(product_ids))
    cursor = conn.cursor()
    try:
        cursor.execute(f"""
            SELECT pi.producto_id, p.sku, p.marca
            FROM productos_intercambiables pi
            INNER JOIN productos p ON p.id = pi.producto_intercambiable_id
            WHERE pi.producto_id IN ({placeholders})
            ORDER BY pi.producto_id, p.marca, p.sku
        """, product_ids)
    except sqlite3.OperationalError:
        return {}

    result: Dict[int, List[str]] = {}
    for row in cursor.fetchall():
        pid = row['producto_id']
        line = f"{row['marca']} - {row['sku']}"
        result.setdefault(pid, []).append(line)
    return result


# ---------------------------------------------------------------------------
# Shared: Download images concurrently
# ---------------------------------------------------------------------------

async def _download_image(client: httpx.AsyncClient, sku: str) -> Optional[bytes]:
    """Download a single product image. Returns bytes or None on failure."""
    url = f"{IMAGE_SERVER_BASE}/{sku}.jpg"
    try:
        resp = await client.get(url)
        if resp.status_code == 200:
            ct = resp.headers.get("content-type", "")
            if ct.startswith("image/") or len(resp.content) > 1000:
                return resp.content
    except (httpx.TimeoutException, httpx.RequestError):
        pass
    return None


async def _download_images_bulk(skus: List[str]) -> Dict[str, bytes]:
    """Download images for a list of SKUs concurrently. Returns {sku: image_bytes}."""
    sem = asyncio.Semaphore(IMAGE_CONCURRENCY)
    results: Dict[str, bytes] = {}

    async def _fetch(sku: str):
        async with sem:
            data = await _download_image(client, sku)
            if data:
                results[sku] = data

    async with httpx.AsyncClient(follow_redirects=True, timeout=IMAGE_DOWNLOAD_TIMEOUT) as client:
        tasks = [_fetch(sku) for sku in skus]
        await asyncio.gather(*tasks)

    return results


def _resize_to_thumbnail(image_bytes: bytes, size: tuple) -> bytes:
    """Resize image to thumbnail PNG bytes."""
    img = PILImage.open(io.BytesIO(image_bytes))
    img.thumbnail(size, PILImage.LANCZOS)
    if img.mode != 'RGB':
        img = img.convert('RGB')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Common filter params (FastAPI dependency)
# ---------------------------------------------------------------------------

def _common_filter_params():
    """Returns the common Query parameters signature. Used in both endpoints."""
    pass  # We define params inline in each endpoint


# ---------------------------------------------------------------------------
# Shared: Fetch products + enrichment
# ---------------------------------------------------------------------------

def _fetch_export_data(
    departamento, marca, marca_vehiculo, modelo_vehiculo, año, motor,
    tipo_producto, grupo_producto, con_inventario, solo_nuevos,
    ancho_llanta, relacion_llanta, diametro_llanta, tipo_llanta, capas_llanta,
    viscosidad, tipo_aceite, presentacion,
    grupo_bci, capacidad_cca, tamano_acumulador,
    q,
) -> List[Dict[str, Any]]:
    """Fetch products + compatibilidades + intercambiables. Raises HTTPException."""

    sql, params = _build_filtered_query(
        departamento, marca, marca_vehiculo, modelo_vehiculo, año, motor,
        tipo_producto, grupo_producto, con_inventario, solo_nuevos,
        ancho_llanta, relacion_llanta, diametro_llanta, tipo_llanta, capas_llanta,
        viscosidad, tipo_aceite, presentacion,
        grupo_bci, capacidad_cca, tamano_acumulador,
        q,
    )

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(sql, params)
        rows = cursor.fetchall()

        if len(rows) == 0:
            raise HTTPException(status_code=404, detail="No se encontraron productos con los filtros aplicados.")
        if len(rows) > MAX_PRODUCTOS_EXPORT:
            raise HTTPException(
                status_code=400,
                detail=f"Demasiados productos ({len(rows)}). El máximo para exportar es {MAX_PRODUCTOS_EXPORT}. Aplique más filtros."
            )

        products = []
        product_ids = []
        for row in rows:
            products.append(dict(row))
            product_ids.append(row['id'])

        compats = _fetch_compatibilidades_bulk(conn, product_ids)
        intercambs = _fetch_intercambiables_bulk(conn, product_ids)

    for p in products:
        p['compatibilidades_list'] = compats.get(p['id'], [])
        p['intercambiables_list'] = intercambs.get(p['id'], [])

    return products


# ---------------------------------------------------------------------------
# Endpoint: Excel
# ---------------------------------------------------------------------------

@router.get("/excel")
async def exportar_excel(
    departamento: Optional[str] = Query(None),
    marca: Optional[str] = Query(None),
    marca_vehiculo: Optional[str] = Query(None),
    modelo_vehiculo: Optional[str] = Query(None),
    año: Optional[int] = Query(None),
    motor: Optional[str] = Query(None),
    tipo_producto: Optional[str] = Query(None),
    grupo_producto: Optional[str] = Query(None),
    con_inventario: Optional[bool] = Query(None),
    solo_nuevos: bool = Query(False),
    ancho_llanta: Optional[str] = Query(None),
    relacion_llanta: Optional[str] = Query(None),
    diametro_llanta: Optional[str] = Query(None),
    tipo_llanta: Optional[str] = Query(None),
    capas_llanta: Optional[str] = Query(None),
    viscosidad: Optional[str] = Query(None),
    tipo_aceite: Optional[str] = Query(None),
    presentacion: Optional[str] = Query(None),
    grupo_bci: Optional[str] = Query(None),
    capacidad_cca: Optional[str] = Query(None),
    tamano_acumulador: Optional[str] = Query(None),
    q: Optional[str] = Query(None, max_length=200),
):
    """Exporta productos filtrados a Excel con imágenes embebidas."""
    products = _fetch_export_data(
        departamento, marca, marca_vehiculo, modelo_vehiculo, año, motor,
        tipo_producto, grupo_producto, con_inventario, solo_nuevos,
        ancho_llanta, relacion_llanta, diametro_llanta, tipo_llanta, capas_llanta,
        viscosidad, tipo_aceite, presentacion,
        grupo_bci, capacidad_cca, tamano_acumulador,
        q,
    )

    # Download images
    skus = [p['sku'] for p in products]
    image_data = await _download_images_bulk(skus)

    # Resize to thumbnails
    thumbnails: Dict[str, bytes] = {}
    for sku, raw in image_data.items():
        try:
            thumbnails[sku] = _resize_to_thumbnail(raw, THUMB_SIZE_EXCEL)
        except Exception:
            pass

    # Build Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo RELUVSA"

    # Header style
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    headers = ["Nuevo", "SKU", "Imagen", "Grupo Producto", "Compatibilidades",
               "Precio Público", "Precio Mayoreo", "Intercambiables"]

    col_widths = [8, 18, 12, 22, 45, 16, 16, 35]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # Data rows
    wrap_alignment = Alignment(vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    price_format = '$#,##0.00'

    for row_idx, product in enumerate(products, 2):
        # Set row height for images
        ws.row_dimensions[row_idx].height = 65

        # Col A: Nuevo
        cell = ws.cell(row=row_idx, column=1, value="Sí" if product['es_nuevo'] else "No")
        cell.alignment = center_alignment
        cell.border = thin_border

        # Col B: SKU
        cell = ws.cell(row=row_idx, column=2, value=product['sku'])
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border

        # Col C: Imagen (embedded)
        cell = ws.cell(row=row_idx, column=3)
        cell.border = thin_border
        sku = product['sku']
        if sku in thumbnails:
            try:
                img_stream = io.BytesIO(thumbnails[sku])
                xl_img = XlImage(img_stream)
                xl_img.width = 75
                xl_img.height = 75
                anchor = f"C{row_idx}"
                ws.add_image(xl_img, anchor)
            except Exception:
                pass

        # Col D: Grupo Producto
        cell = ws.cell(row=row_idx, column=4, value=product.get('grupo_producto') or '')
        cell.alignment = wrap_alignment
        cell.border = thin_border

        # Col E: Compatibilidades (multiline)
        compat_lines = product.get('compatibilidades_list', [])
        cell = ws.cell(row=row_idx, column=5, value="\n".join(compat_lines))
        cell.alignment = wrap_alignment
        cell.border = thin_border

        # Col F: Precio Público
        cell = ws.cell(row=row_idx, column=6, value=product.get('precio_publico') or 0)
        cell.number_format = price_format
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = thin_border

        # Col G: Precio Mayoreo
        cell = ws.cell(row=row_idx, column=7, value=product.get('precio_mayoreo') or 0)
        cell.number_format = price_format
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = thin_border

        # Col H: Intercambiables (multiline)
        intercamb_lines = product.get('intercambiables_list', [])
        cell = ws.cell(row=row_idx, column=8, value="\n".join(intercamb_lines))
        cell.alignment = wrap_alignment
        cell.border = thin_border

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=catalogo_reluvsa.xlsx"
        }
    )


# ---------------------------------------------------------------------------
# Endpoint: PDF
# ---------------------------------------------------------------------------

@router.get("/pdf")
async def exportar_pdf(
    departamento: Optional[str] = Query(None),
    marca: Optional[str] = Query(None),
    marca_vehiculo: Optional[str] = Query(None),
    modelo_vehiculo: Optional[str] = Query(None),
    año: Optional[int] = Query(None),
    motor: Optional[str] = Query(None),
    tipo_producto: Optional[str] = Query(None),
    grupo_producto: Optional[str] = Query(None),
    con_inventario: Optional[bool] = Query(None),
    solo_nuevos: bool = Query(False),
    ancho_llanta: Optional[str] = Query(None),
    relacion_llanta: Optional[str] = Query(None),
    diametro_llanta: Optional[str] = Query(None),
    tipo_llanta: Optional[str] = Query(None),
    capas_llanta: Optional[str] = Query(None),
    viscosidad: Optional[str] = Query(None),
    tipo_aceite: Optional[str] = Query(None),
    presentacion: Optional[str] = Query(None),
    grupo_bci: Optional[str] = Query(None),
    capacidad_cca: Optional[str] = Query(None),
    tamano_acumulador: Optional[str] = Query(None),
    q: Optional[str] = Query(None, max_length=200),
):
    """Exporta productos filtrados a PDF con imágenes embebidas."""
    products = _fetch_export_data(
        departamento, marca, marca_vehiculo, modelo_vehiculo, año, motor,
        tipo_producto, grupo_producto, con_inventario, solo_nuevos,
        ancho_llanta, relacion_llanta, diametro_llanta, tipo_llanta, capas_llanta,
        viscosidad, tipo_aceite, presentacion,
        grupo_bci, capacidad_cca, tamano_acumulador,
        q,
    )

    # Download images
    skus = [p['sku'] for p in products]
    image_data = await _download_images_bulk(skus)

    # Resize to PDF thumbnails
    thumbnails: Dict[str, bytes] = {}
    for sku, raw in image_data.items():
        try:
            thumbnails[sku] = _resize_to_thumbnail(raw, THUMB_SIZE_PDF)
        except Exception:
            pass

    # Build PDF
    buffer = io.BytesIO()
    page_w, page_h = landscape(letter)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    elements = []

    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleReluvsa',
        parent=styles['Title'],
        fontSize=16,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        'SubtitleReluvsa',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#666666'),
        spaceAfter=12,
    )

    elements.append(Paragraph("Catálogo RELUVSA", title_style))
    elements.append(Paragraph(f"{len(products)} productos", subtitle_style))

    # Cell styles for table content
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
    )
    cell_style_small = ParagraphStyle(
        'CellStyleSmall',
        parent=styles['Normal'],
        fontSize=6.5,
        leading=8,
        textColor=colors.HexColor('#333333'),
    )

    # Table header
    header_row = [
        Paragraph("<b>Imagen</b>", cell_style),
        Paragraph("<b>Marca</b>", cell_style),
        Paragraph("<b>No. Parte</b>", cell_style),
        Paragraph("<b>Grupo Producto</b>", cell_style),
        Paragraph("<b>Compatibilidades</b>", cell_style),
        Paragraph("<b>P. Público</b>", cell_style),
        Paragraph("<b>P. Mayoreo</b>", cell_style),
        Paragraph("<b>Intercambiables</b>", cell_style),
    ]

    table_data = [header_row]

    for product in products:
        # Image (use sku for download key)
        sku = product['sku']
        if sku in thumbnails:
            try:
                img_io = io.BytesIO(thumbnails[sku])
                rl_img = RLImage(img_io, width=50, height=50)
            except Exception:
                rl_img = ""
        else:
            rl_img = ""

        # Marca
        marca_prod = product.get('marca') or ''

        # Número de parte (primer sku alterno, fallback a sku/clave)
        num_parte = sku
        skus_alt_raw = product.get('skus_alternos')
        if skus_alt_raw:
            try:
                skus_alt = json.loads(skus_alt_raw) if isinstance(skus_alt_raw, str) else skus_alt_raw
                if skus_alt and len(skus_alt) > 0:
                    num_parte = skus_alt[0]
            except (ValueError, TypeError):
                pass

        # Grupo Producto
        grupo = product.get('grupo_producto') or ''

        # Compatibilidades (limit to 5 + "... y X más")
        compat_all = product.get('compatibilidades_list', [])
        if len(compat_all) > 5:
            compat_display = compat_all[:5]
            compat_display.append(f"... y {len(compat_all) - 5} más")
        else:
            compat_display = compat_all
        compat_text = "<br/>".join(compat_display) if compat_display else ""

        # Precios
        pp = product.get('precio_publico') or 0
        pm = product.get('precio_mayoreo') or 0
        precio_pub = f"${pp:,.2f}"
        precio_may = f"${pm:,.2f}"

        # Intercambiables (limit to 5)
        intercamb_all = product.get('intercambiables_list', [])
        if len(intercamb_all) > 5:
            intercamb_display = intercamb_all[:5]
            intercamb_display.append(f"... y {len(intercamb_all) - 5} más")
        else:
            intercamb_display = intercamb_all
        intercamb_text = "<br/>".join(intercamb_display) if intercamb_display else ""

        row = [
            rl_img,
            Paragraph(marca_prod, cell_style),
            Paragraph(num_parte, cell_style),
            Paragraph(grupo, cell_style),
            Paragraph(compat_text, cell_style_small),
            Paragraph(precio_pub, cell_style),
            Paragraph(precio_may, cell_style),
            Paragraph(intercamb_text, cell_style_small),
        ]
        table_data.append(row)

    # Column widths (landscape letter ~10.5" usable after margins)
    available_width = page_w - 30 * mm  # left + right margins
    col_widths = [
        55,   # Imagen
        60,   # Marca
        65,   # No. Parte
        80,   # Grupo
        140,  # Compatibilidades
        55,   # P. Público
        55,   # P. Mayoreo
        available_width - 55 - 60 - 65 - 80 - 140 - 55 - 55,  # Intercambiables
    ]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Table style
    style_commands = [
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFD700')),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),

        # All cells
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]

    # Alternating row colors
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style_commands.append(
                ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f5f5f5'))
            )

    table.setStyle(TableStyle(style_commands))
    elements.append(table)

    # Footer: Precios NETOS
    footer_style = ParagraphStyle(
        'FooterReluvsa',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#333333'),
        spaceBefore=12,
        alignment=1,  # CENTER
    )
    elements.append(Spacer(1, 8))
    elements.append(Paragraph("<b>* Los precios mostrados son NETOS (incluyen IVA)</b>", footer_style))

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=catalogo_reluvsa.pdf"
        }
    )
