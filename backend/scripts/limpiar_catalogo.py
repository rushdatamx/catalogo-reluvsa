"""
Script de limpieza del catálogo RELUVSA.
Elimina ~13,700 productos inactivos (inventario 0 + sin ventas recientes).

Uso:
    cd backend
    python3 scripts/limpiar_catalogo.py
"""

import sqlite3
import shutil
import sys
from pathlib import Path
from datetime import datetime, timedelta

# --- Paths ---
DB_PATH = Path(__file__).parent.parent.parent / "data" / "catalogo.db"
EXCEL_PATH = Path(__file__).parent.parent.parent / "data" / "limpieza-catalogo.xlsx"
BACKUP_DIR = Path(__file__).parent.parent.parent / "data"

# --- Constantes ---
FECHA_CORTE = datetime.now() - timedelta(days=5 * 365)  # 5 años atrás
BATCH_SIZE = 500  # SQLite tiene límite ~999 variables en IN clause


def normalizar_sku(sku: str) -> str:
    """Normaliza SKU: strip + quitar ceros iniciales en numéricos."""
    sku = sku.strip()
    if sku.isdigit():
        return sku.lstrip('0') or '0'
    return sku


def cargar_excel() -> dict:
    """
    Carga el Excel del cliente y retorna diccionario:
    {sku_normalizado: {'inventario': int, 'ultima_venta': datetime|None}}
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl no instalado. Ejecuta: pip install openpyxl")
        sys.exit(1)

    print(f"Cargando Excel: {EXCEL_PATH}")
    if not EXCEL_PATH.exists():
        print(f"ERROR: No se encontró el archivo {EXCEL_PATH}")
        sys.exit(1)

    wb = load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    ws = wb.active

    datos = {}
    filas_procesadas = 0
    filas_sin_clave = 0

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Col A (0) = Clave, Col G (6) = Ultima Venta, Col V (21) = Total Almacenes
        clave_raw = row[0]
        if clave_raw is None:
            filas_sin_clave += 1
            continue

        sku = normalizar_sku(str(clave_raw))

        # Total Almacenes (col V = index 21)
        inventario_raw = row[21]
        try:
            inventario = int(float(inventario_raw)) if inventario_raw is not None else 0
        except (ValueError, TypeError):
            inventario = 0

        # Ultima Venta (col G = index 6)
        ultima_venta_raw = row[6]
        ultima_venta = None
        if ultima_venta_raw is not None:
            if isinstance(ultima_venta_raw, datetime):
                ultima_venta = ultima_venta_raw
            else:
                # Intentar parsear string
                for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y'):
                    try:
                        ultima_venta = datetime.strptime(str(ultima_venta_raw), fmt)
                        break
                    except ValueError:
                        continue

        datos[sku] = {
            'inventario': inventario,
            'ultima_venta': ultima_venta,
        }
        filas_procesadas += 1

    wb.close()

    print(f"  SKUs cargados del Excel: {len(datos):,}")
    if filas_sin_clave > 0:
        print(f"  Filas sin clave (ignoradas): {filas_sin_clave}")

    return datos


def obtener_productos_db(conn) -> list:
    """Retorna todos los productos de la DB: [(id, sku), ...]"""
    cursor = conn.execute("SELECT id, sku FROM productos")
    return cursor.fetchall()


def clasificar_productos(productos_db: list, datos_excel: dict) -> dict:
    """
    Clasifica cada producto en conservar o borrar según las reglas de decisión.
    Retorna dict con listas de IDs y estadísticas.
    """
    conservar = []
    borrar = []

    # Estadísticas por categoría
    stats = {
        'con_inventario': 0,
        'sin_inv_venta_reciente': 0,
        'sin_inv_venta_vieja': 0,
        'sin_inv_sin_venta': 0,
        'no_en_excel': 0,
    }

    for prod_id, sku in productos_db:
        sku_norm = normalizar_sku(sku)

        if sku_norm not in datos_excel:
            # SKU en DB pero NO en Excel → BORRAR
            borrar.append(prod_id)
            stats['no_en_excel'] += 1
            continue

        info = datos_excel[sku_norm]

        if info['inventario'] > 0:
            # Tiene inventario → CONSERVAR
            conservar.append(prod_id)
            stats['con_inventario'] += 1
        elif info['ultima_venta'] is not None and info['ultima_venta'] >= FECHA_CORTE:
            # Sin inventario pero venta reciente (< 5 años) → CONSERVAR
            conservar.append(prod_id)
            stats['sin_inv_venta_reciente'] += 1
        elif info['ultima_venta'] is not None and info['ultima_venta'] < FECHA_CORTE:
            # Sin inventario y venta vieja (>= 5 años) → BORRAR
            borrar.append(prod_id)
            stats['sin_inv_venta_vieja'] += 1
        else:
            # Sin inventario y sin fecha de venta → BORRAR
            borrar.append(prod_id)
            stats['sin_inv_sin_venta'] += 1

    return {
        'conservar': conservar,
        'borrar': borrar,
        'stats': stats,
    }


def obtener_departamentos_afectados(conn, ids_borrar: list) -> dict:
    """Retorna conteo de productos a borrar por departamento."""
    deptos = {}
    for i in range(0, len(ids_borrar), BATCH_SIZE):
        batch = ids_borrar[i:i + BATCH_SIZE]
        placeholders = ','.join('?' * len(batch))
        cursor = conn.execute(
            f"SELECT departamento, COUNT(*) FROM productos WHERE id IN ({placeholders}) GROUP BY departamento",
            batch
        )
        for row in cursor:
            depto = row[0] or '(Sin departamento)'
            deptos[depto] = deptos.get(depto, 0) + row[1]
    return dict(sorted(deptos.items(), key=lambda x: -x[1]))


def conteos_tablas(conn) -> dict:
    """Retorna conteos de todas las tablas relevantes."""
    tablas = ['productos', 'compatibilidades', 'inventario',
              'caracteristicas_producto', 'productos_intercambiables',
              'especificaciones_manuales']
    conteos = {}
    for tabla in tablas:
        cursor = conn.execute(f"SELECT COUNT(*) FROM {tabla}")
        conteos[tabla] = cursor.fetchone()[0]
    return conteos


def mostrar_reporte_pre(stats: dict, ids_borrar: list, ids_conservar: list, deptos: dict):
    """Muestra reporte detallado antes de borrar."""
    print(f"\n{'=' * 70}")
    print("REPORTE PRE-BORRADO")
    print(f"{'=' * 70}")

    print(f"\n--- Clasificación de productos ---")
    print(f"  CONSERVAR:")
    print(f"    Con inventario > 0:              {stats['con_inventario']:>8,}")
    print(f"    Sin inv, venta < 5 años:         {stats['sin_inv_venta_reciente']:>8,}")
    print(f"  BORRAR:")
    print(f"    Sin inv, venta >= 5 años:        {stats['sin_inv_venta_vieja']:>8,}")
    print(f"    Sin inv, sin fecha de venta:     {stats['sin_inv_sin_venta']:>8,}")
    print(f"    SKU en DB pero no en Excel:      {stats['no_en_excel']:>8,}")

    total_conservar = len(ids_conservar)
    total_borrar = len(ids_borrar)
    print(f"\n  Total CONSERVAR: {total_conservar:>8,}")
    print(f"  Total BORRAR:    {total_borrar:>8,}")
    print(f"  {'─' * 35}")
    print(f"  Total productos: {total_conservar + total_borrar:>8,}")

    print(f"\n--- Departamentos afectados (productos a borrar) ---")
    for depto, count in deptos.items():
        print(f"    {depto:<40} {count:>6,}")

    print(f"\n{'=' * 70}")


def borrar_en_cascada(conn, ids_borrar: list):
    """Borra productos y sus relaciones en cascada, por batches."""
    tablas_hijas = [
        ('compatibilidades', 'producto_id'),
        ('inventario', 'producto_id'),
        ('caracteristicas_producto', 'producto_id'),
        ('especificaciones_manuales', 'producto_id'),
    ]

    total_borrados = {}

    for tabla, columna in tablas_hijas:
        count = 0
        for i in range(0, len(ids_borrar), BATCH_SIZE):
            batch = ids_borrar[i:i + BATCH_SIZE]
            placeholders = ','.join('?' * len(batch))
            cursor = conn.execute(
                f"DELETE FROM {tabla} WHERE {columna} IN ({placeholders})",
                batch
            )
            count += cursor.rowcount
        total_borrados[tabla] = count
        print(f"  {tabla}: {count:,} registros eliminados")

    # productos_intercambiables: borrar en ambas direcciones
    count_inter = 0
    for i in range(0, len(ids_borrar), BATCH_SIZE):
        batch = ids_borrar[i:i + BATCH_SIZE]
        placeholders = ','.join('?' * len(batch))
        cursor = conn.execute(
            f"DELETE FROM productos_intercambiables WHERE producto_id IN ({placeholders}) OR producto_intercambiable_id IN ({placeholders})",
            batch + batch
        )
        count_inter += cursor.rowcount
    total_borrados['productos_intercambiables'] = count_inter
    print(f"  productos_intercambiables: {count_inter:,} registros eliminados")

    # Finalmente borrar de productos
    count_prod = 0
    for i in range(0, len(ids_borrar), BATCH_SIZE):
        batch = ids_borrar[i:i + BATCH_SIZE]
        placeholders = ','.join('?' * len(batch))
        cursor = conn.execute(
            f"DELETE FROM productos WHERE id IN ({placeholders})",
            batch
        )
        count_prod += cursor.rowcount
    total_borrados['productos'] = count_prod
    print(f"  productos: {count_prod:,} registros eliminados")

    return total_borrados


def limpiar_intercambiables_huerfanos(conn):
    """Elimina referencias a productos que ya no existen en productos_intercambiables."""
    cursor = conn.execute(
        "DELETE FROM productos_intercambiables WHERE producto_intercambiable_id NOT IN (SELECT id FROM productos)"
    )
    huerfanos = cursor.rowcount
    if huerfanos > 0:
        print(f"  Intercambiables huérfanos limpiados: {huerfanos:,}")
    else:
        print(f"  Sin intercambiables huérfanos")
    return huerfanos


def main():
    print(f"{'=' * 70}")
    print("LIMPIEZA DEL CATÁLOGO RELUVSA")
    print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Fecha de corte (5 años): {FECHA_CORTE.strftime('%Y-%m-%d')}")
    print(f"{'=' * 70}")

    # Verificar DB
    if not DB_PATH.exists():
        print(f"ERROR: No se encontró la base de datos en {DB_PATH}")
        sys.exit(1)

    db_size_antes = DB_PATH.stat().st_size

    # --- PASO 1: Backup ---
    print(f"\n--- Paso 1: Backup ---")
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = BACKUP_DIR / f"catalogo_backup_pre_limpieza_{timestamp}.db"
    shutil.copy2(str(DB_PATH), str(backup_path))

    # Verificar backup
    if backup_path.exists() and backup_path.stat().st_size == db_size_antes:
        print(f"  Backup creado: {backup_path.name}")
        print(f"  Tamaño: {db_size_antes / (1024*1024):.1f} MB")
    else:
        print("ERROR: El backup no se creó correctamente. Abortando.")
        sys.exit(1)

    # --- PASO 2: Cargar Excel ---
    print(f"\n--- Paso 2: Cargar Excel ---")
    datos_excel = cargar_excel()

    # --- PASO 3: Clasificar productos ---
    print(f"\n--- Paso 3: Clasificar productos ---")
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row

    productos_db = obtener_productos_db(conn)
    print(f"  Productos en DB: {len(productos_db):,}")

    resultado = clasificar_productos(productos_db, datos_excel)
    ids_conservar = resultado['conservar']
    ids_borrar = resultado['borrar']
    stats = resultado['stats']

    # --- PASO 4: Reporte pre-borrado ---
    deptos = obtener_departamentos_afectados(conn, ids_borrar)
    conteos_antes = conteos_tablas(conn)

    mostrar_reporte_pre(stats, ids_borrar, ids_conservar, deptos)

    print("\nConteos ANTES del borrado:")
    for tabla, count in conteos_antes.items():
        print(f"    {tabla:<35} {count:>8,}")

    # --- Confirmación interactiva ---
    print(f"\n{'!' * 70}")
    print(f"  SE VAN A BORRAR {len(ids_borrar):,} PRODUCTOS Y SUS DATOS ASOCIADOS")
    print(f"  Esta acción es IRREVERSIBLE (excepto restaurando el backup)")
    print(f"{'!' * 70}")
    print(f"\n  Backup disponible en: {backup_path.name}")

    if '--confirmar' in sys.argv:
        print("\n  Flag --confirmar detectado. Procediendo...")
    else:
        confirmacion = input('\n  Escribe CONFIRMAR para proceder: ')
        if confirmacion.strip() != 'CONFIRMAR':
            print("\n  Operación cancelada por el usuario.")
            conn.close()
            sys.exit(0)

    # --- PASO 5: Borrar en cascada ---
    print(f"\n--- Paso 5: Borrar en cascada ---")
    try:
        borrar_en_cascada(conn, ids_borrar)

        # --- PASO 6: Limpiar intercambiables huérfanos ---
        print(f"\n--- Paso 6: Limpiar intercambiables huérfanos ---")
        limpiar_intercambiables_huerfanos(conn)

        conn.commit()
        print("\n  COMMIT exitoso.")
    except Exception as e:
        conn.rollback()
        print(f"\n  ERROR: {e}")
        print("  ROLLBACK ejecutado. La base de datos no fue modificada.")
        conn.close()
        sys.exit(1)

    conn.close()

    # --- PASO 7: VACUUM ---
    print(f"\n--- Paso 7: VACUUM ---")
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("VACUUM")
    conn.close()
    print("  VACUUM completado.")

    # --- PASO 8: Reporte post-borrado ---
    db_size_despues = DB_PATH.stat().st_size

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conteos_despues = conteos_tablas(conn)
    conn.close()

    print(f"\n{'=' * 70}")
    print("REPORTE POST-BORRADO")
    print(f"{'=' * 70}")

    print(f"\n{'Tabla':<35} {'Antes':>10} {'Después':>10} {'Borrados':>10}")
    print(f"{'─' * 65}")
    for tabla in conteos_antes:
        antes = conteos_antes[tabla]
        despues = conteos_despues[tabla]
        diff = antes - despues
        print(f"{tabla:<35} {antes:>10,} {despues:>10,} {diff:>10,}")

    print(f"\nTamaño de la base de datos:")
    print(f"  Antes:   {db_size_antes / (1024*1024):.1f} MB")
    print(f"  Después: {db_size_despues / (1024*1024):.1f} MB")
    print(f"  Reducción: {(db_size_antes - db_size_despues) / (1024*1024):.1f} MB ({(1 - db_size_despues/db_size_antes)*100:.1f}%)")

    print(f"\n{'=' * 70}")
    print("LIMPIEZA COMPLETADA")
    print(f"Backup disponible en: {backup_path.name}")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
